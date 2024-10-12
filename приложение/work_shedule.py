import tkinter as tk
from tkinter import ttk
import calendar
from datetime import datetime, timedelta
import openpyxl
import mysql.connector
import os


def work_shedule(canvas):
    def connect_to_db():
        return mysql.connector.connect(
            host="localhost",  # Замените на ваш хост
            user="root",  # Замените на ваше имя пользователя
            password="",  # Замените на ваш пароль
            database="sklad"  # Название базы данных
        )

    # Функция для загрузки данных из таблицы 'заказ'
    def load_data_from_db(query):
        conn = connect_to_db()
        cursor = conn.cursor()

        # Выполняем SQL-запрос для получения данных из таблицы 'заказ'
        cursor.execute(query)

        # Получаем все строки из результата запроса
        rows = cursor.fetchall()

        # Закрываем соединение
        conn.close()

        return rows

    def find_column_by_date(file_path, date):
        # Загружаем файл Excel
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # Берём активный лист (по умолчанию первый)

        # Преобразуем дату в строку (если она не в строковом формате)
        date_str = str(date)
        # Перебираем первую строку (заголовки) для поиска нужной даты
        column_index = None
        for col in range(1, sheet.max_column + 1):
            cell_value = str(sheet.cell(row=1, column=col).value)
            if cell_value == date_str:
                column_index = col
                break

        # Если дата не найдена, возвращаем None
        if column_index is None:
            return None

        # Извлекаем весь столбец с найденной датой
        column_data = []
        for row in range(2, sheet.max_row + 1):  # Начинаем со второй строки, пропуская заголовок
            cell_value = sheet.cell(row=row, column=column_index).value
            column_data.append(cell_value)

        return column_data

    # Функция для обновления меток с датами
    def update_label():
        # Получаем начало и конец месяца
        start_date = datetime(current_date.year, current_date.month, 1)  # Начало месяца
        _, last_day = calendar.monthrange(current_date.year, current_date.month)
        end_date = datetime(current_date.year, current_date.month, last_day)  # Конец месяца

        # Обновляем текст метки
        month_label.config(text=f"({months_ru[current_date.month - 1]} {current_date.year})")
        range_label.config(text=f"({start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')})")

        day_count = 1
        # Определяем, с какого дня недели начинается месяц
        first_day_weekday = start_date.weekday()

        # Обновляем кнопки для календаря
        for row in range(ROWS):
            # Сравниваем даты
            if start_date <= datetime.today() <= end_date and current_date.day == day_count and str(day_count) == str(
                    datetime.today().day):
                buttons[row][0].config(background='grey')  # Подсвечиваем текущий день
            else:
                buttons[row][0].config(background='SystemButtonFace')
            if row < first_day_weekday:
                buttons[row][0].config(text='', state='disabled')  # Отключаем кнопки до начала месяца
            else:
                buttons[row][0].config(text=str(day_count), state='normal')
                day_count += 1

        for col in range(1, COLUMNS):
            for row in range(ROWS):
                if start_date <= datetime.today() <= end_date and current_date.day == day_count and str(
                        day_count) == str(
                    datetime.today().day):
                    buttons[row][col].config(background='grey')  # Подсвечиваем текущий день
                else:
                    buttons[row][col].config(background='SystemButtonFace')

                if day_count <= last_day:
                    buttons[row][col].config(text=str(day_count), state='normal')  # Активируем кнопки в пределах месяца
                else:
                    buttons[row][col].config(text='', state='disabled')  # Деактивируем кнопки после конца месяца

                day_count += 1

    # Функция для увеличения месяца на 1
    def next_month():
        global current_date
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)
        update_label()

    # Функция для уменьшения месяца на 1
    def prev_month():
        global current_date
        if current_date.month == 1:
            current_date = current_date.replace(year=current_date.year - 1, month=12)
        else:
            current_date = current_date.replace(month=current_date.month - 1)
        update_label()

    # Функция, которая вызывается при нажатии на кнопку
    def button_clicked(row, col):

        data_for_exel = str(current_date)[:8] + [
            buttons[row][col]['text'] if int(buttons[row][col]['text']) >= 10 else '0' + buttons[row][col]['text']][
            0] + ' 00:00:00'
        column_data = find_column_by_date(file_path_exel, data_for_exel)
        data_employees = load_data_from_db(
            f'select id_сотрудник,логин,фамилия,имя,отчество from сотрудник where специальность="{table_dropdown_spec.get()}";')
        if column_data is not None:
            id_date = []
            for i in range(len(column_data)):
                if column_data[i] == table_dropdown_smena.get()[0]:
                    id_date.append(i)
            data_employees = [i for i in data_employees if i[0] in id_date]

            # Очищаем текущее содержимое Treeview
            for row in worked_employee_treeview.get_children():
                worked_employee_treeview.delete(row)
            # Заполняем Treeview новыми данными
            for row in data_employees:
                worked_employee_treeview.insert('', tk.END, values=row)

            label_data['text'] = f'Дата: {data_for_exel[:-9]}'
        else:
            # Очищаем текущее содержимое Treeview
            for row in worked_employee_treeview.get_children():
                worked_employee_treeview.delete(row)
        # Очищаем текущее содержимое Treeview
        for row in treeview_pik.get_children():
            treeview_pik.delete(row)

        quary_data = ['', '', '']

        if table_dropdown_smena.get() == 'дневная':
            quary_data[1] = f'"{data_for_exel[:10]} 9:00:00" and "{data_for_exel[:10]} 21:00:00"'
        else:
            # Преобразуем строку в объект datetime
            date_obj = datetime.strptime(data_for_exel[:10], "%Y-%m-%d")
            # Прибавляем один день
            new_date_obj = date_obj + timedelta(days=1)
            # Преобразуем обратно в строку, если нужно
            new_date_str = new_date_obj.strftime("%Y-%m-%d")

            quary_data[1] = f'"{data_for_exel[:10]} 21:00:00" and "{new_date_str} 9:00:00"'
        if table_dropdown_spec.get() != "Менеджер":
            if table_dropdown_spec.get() == "Комплектовщик ячеек":
                quary = f'''
    select 
    s2.логин, 
    count(*) as "количество пиков"  
    from товар z
    -- Соединение с комплектовщиком
    JOIN `комплектовщик ячеек` k ON z.`комплектовщик ячеек_id_комплектовщик_ячеек` = k.id_комплектовщик_ячеек
    JOIN сотрудник s2 ON k.`сотрудник_id_сотрудник` = s2.id_сотрудник
    where 
    z.время_распределения BETWEEN {quary_data[1]} 
    group by z.`комплектовщик ячеек_id_комплектовщик_ячеек`
    order by count(*) desc; 
    '''
            elif table_dropdown_spec.get() == "Принимающий поставку":
                quary = f'''
    select 
    s2.логин,
    count(*)
    from товар z
    -- соединение с принимающим поставку
    join поставка s1 on z.поставка_id_поставка=s1.id_поставка
    JOIN `принимающий поставку` k ON s1.`принимающий поставку_id_принимающий_поставку` = k.`id_принимающий_поставку`
    JOIN сотрудник s2 ON k.`сотрудник_id_сотрудник` = s2.id_сотрудник
    where s1.`дата приёма` BETWEEN {quary_data[1]}
    group by s1.`принимающий поставку_id_принимающий_поставку`
    order by count(*) desc           
                '''
            elif table_dropdown_spec.get() == "Водитель":
                quary = f'''
    SELECT s.логин, COUNT(*) AS общее_количество
    FROM сотрудник s
    JOIN водитель v ON s.id_сотрудник = v.сотрудник_id_сотрудник
    JOIN (
        SELECT водитель_id_водитель
        FROM поставка
        where `дата приёма` BETWEEN {quary_data[1]}
        UNION ALL
        SELECT водитель_id_водитель
        FROM заказ
        where время_отправки BETWEEN {quary_data[1]}
    ) AS объединённый_запрос ON v.id_водитель = объединённый_запрос.водитель_id_водитель
    GROUP BY s.логин
    order by count(*) desc  ;
    '''
            else:
                if table_dropdown_spec.get() == "Комплектовщик":
                    quary_data[0] = 'JOIN комплектовщик k ON z.`комплектовщик_id_комплектовщик` = k.id_комплектовщик'
                    quary_data[2] = "group by z.комплектовщик_id_комплектовщик"
                elif table_dropdown_spec.get() == "Проверяющий комплектовку":
                    quary_data[
                        0] = 'JOIN `проверяющий комплектовку` k ON z.`проверяющий комплектовку_id_проверяющий_комплектовку` = k.id_проверяющий_комплектовку'
                    quary_data[2] = "group by z.`проверяющий комплектовку_id_проверяющий_комплектовку`"
                quary = f'''
    select 
    s2.логин, 
    count(*) as "количество пиков"
    from заказ z
    join `товары в заказе` s1 on z.id_заказ = s1.заказ_id_заказ
    {quary_data[0]}
    JOIN сотрудник s2 ON k.`сотрудник_id_сотрудник` = s2.id_сотрудник
    where
    z.время_на_сборку BETWEEN {quary_data[1]} 
    {quary_data[2]}
    order by count(*) desc;
                '''
            print(quary)
            data = load_data_from_db(quary)
            for item in data:
                treeview_pik.insert("", tk.END, values=item)
        summa = 0
        for child in treeview_pik.get_children():
            summa += int(treeview_pik.item(child, 'values')[1])
        label_itog['text'] = f'Итог: {summa}'

    file_path_exel = 'C:/Users/Пасечниковы/Desktop/sklad/приложение/расписание.xlsx'  # Укажите путь к вашему файлу Excel
    # Заголовки для дней недели
    days_of_week = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

    tk.Label(canvas, text="ГРАФИК РАБОТЫ", font=("Arial", 30)).place(x=30, y=70)

    # Фрейм для сетки
    grid_frame = tk.Frame(canvas, width=GRID_WIDTH, height=GRID_HEIGHT)
    grid_frame.place(x=0, y=150)

    # Создание заголовков для дней недели (слева от таблицы)
    for col in range(ROWS):
        lbl = tk.Label(grid_frame, text=days_of_week[col], width=5, height=2, font=("Arial", 24))
        lbl.grid(row=col, column=0, sticky="nsew")  # Добавляем метки в первый столбец

    # Список для хранения кнопок
    buttons = []

    # Создание сетки из кнопок (с 1 столбца, т.к. 0 занят заголовками дней недели)
    for row in range(ROWS):
        button_row = []  # Список для хранения кнопок в текущем ряду
        for col in range(COLUMNS):
            btn = tk.Button(grid_frame, text=f"", font=("Arial", 24), width=6, height=2,
                            command=lambda r=row, c=col: button_clicked(r, c))
            btn.grid(row=row, column=col + 1, sticky="nsew")  # Сдвигаем на 1 столбец вправо
            button_row.append(btn)  # Добавляем кнопку в текущий ряд
        buttons.append(button_row)  # Добавляем ряд кнопок в общий список

    # Метки для отображения текущего месяца и диапазона дат
    month_label = tk.Label(canvas, font=("Arial", 20))
    month_label.place(x=300, y=900)

    range_label = tk.Label(canvas, font=("Arial", 14))
    range_label.place(x=300, y=940)

    # Кнопки для переключения месяцев
    prev_button = tk.Button(canvas, text="←", command=prev_month, width=10)
    prev_button.place(x=200, y=905)

    next_button = tk.Button(canvas, text="→", command=next_month, width=10)
    next_button.place(x=550, y=905)

    worked_employee_treeview = ttk.Treeview(canvas, columns=('#1', '#2', '#3', '#4', '#5'), show="headings")

    worked_employee_treeview.heading('#1', text='id_сотрудника')
    worked_employee_treeview.heading('#2', text='логин')
    worked_employee_treeview.heading('#3', text='фамилия')
    worked_employee_treeview.heading('#4', text='имя')
    worked_employee_treeview.heading('#5', text='отчество')

    worked_employee_treeview.place(x=880, y=430, height=410)

    label_spec = tk.Label(canvas, text="Должность:", font=("Helvetica", 20))
    label_spec.place(x=900, y=80)

    # Создаем кнопки
    spec_text = [
        "Водитель",
        "Менеджер",
        "Комплектовщик",
        "Проверяющий комплектовку",
        "Комплектовщик ячеек",
        "Принимающий поставку"
    ]

    # Выпадающий список таблиц
    table_dropdown_spec = ttk.Combobox(canvas, state="readonly", font=("Arial", 12))
    table_dropdown_spec.place(x=900, y=150)
    table_dropdown_spec['values'] = spec_text
    table_dropdown_spec.set(spec_text[0])

    label_smena = tk.Label(canvas, text="Смена:", font=("Helvetica", 20))
    label_smena.place(x=900, y=230)

    label_data = tk.Label(canvas, text="Дата:", font=("Helvetica", 20))
    label_data.place(x=900, y=350)

    # Выпадающий список таблиц
    table_dropdown_smena = ttk.Combobox(canvas, state="readonly", font=("Arial", 12))
    table_dropdown_smena.place(x=900, y=300)
    table_dropdown_smena['values'] = ['ночная', 'дневная']
    table_dropdown_smena.set(['ночная'])

    # Создаем кнопку
    open_button = tk.Button(canvas, font=("Arial", 14), text="Открыть Excel файл", command=lambda: os.startfile(
        'C:/Users/Пасечниковы/Desktop/sklad/приложение/расписание.xlsx'))
    open_button.place(x=1600, y=880)

    label_uspev = tk.Label(canvas, text="Успеваемость:", font=("Helvetica", 20))
    label_uspev.place(x=1300, y=50)

    treeview_pik = ttk.Treeview(canvas, columns=('#1', '#2'), show="headings")
    treeview_pik.heading('#1', text='логин')
    treeview_pik.heading('#2', text='количество пиков')
    treeview_pik.place(x=1300, y=110)

    label_itog = tk.Label(canvas, text="Итог: 0", font=("Helvetica", 20))
    label_itog.place(x=1550, y=350)

    update_label()


# Размеры окна
WINDOW_WIDTH = 1920
WINDOW_HEIGHT = 1080

# Размеры сетки (половина экрана)
GRID_WIDTH = WINDOW_WIDTH // 2
GRID_HEIGHT = WINDOW_HEIGHT // 2

# Количество колонок и строк (без учёта дней недели)
COLUMNS = 6
ROWS = 7

# Переменная для хранения текущего года и месяца
current_date = datetime.now()
months_ru = [
    "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
]

if __name__ == '__main__':
    # Основное окно
    root = tk.Tk()
    root.title("Управление заказами")
    # Canvas для отображения выбранного текста
    canvas = tk.Canvas(root, width=1920, height=1080)
    canvas.pack()
    work_shedule(canvas)
    root.mainloop()
